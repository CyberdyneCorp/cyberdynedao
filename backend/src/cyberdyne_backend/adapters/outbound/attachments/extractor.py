"""Multi-format document text extraction (``TextExtractorPort``).

Each supported MIME type maps to a small private parser. The blocking
parse work (pypdf, python-docx, openpyxl, csv) runs in a worker thread so
the event loop isn't stalled. Output is capped so a large document can't
blow up the LLM prompt; unknown types yield "".
"""

from __future__ import annotations

import asyncio
import csv
import io

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

# Hard ceiling on extracted text fed into the prompt.
_MAX_CHARS = 20_000

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class MultiFormatTextExtractor:
    async def extract(self, *, data: bytes, content_type: str) -> str:
        mime = content_type.split(";", 1)[0].strip().lower()
        text = await asyncio.to_thread(self._extract_sync, data, mime)
        return text[:_MAX_CHARS]

    def _extract_sync(self, data: bytes, mime: str) -> str:
        if mime == "application/pdf":
            return self._from_pdf(data)
        if mime == _DOCX_MIME:
            return self._from_docx(data)
        if mime == "text/csv":
            return self._from_csv(data)
        if mime == _XLSX_MIME:
            return self._from_xlsx(data)
        if mime.startswith("text/"):
            return data.decode("utf-8", errors="replace")
        return ""

    def _from_pdf(self, data: bytes) -> str:
        reader = PdfReader(io.BytesIO(data))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n\n".join(pages)

    def _from_docx(self, data: bytes) -> str:
        document = Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs)

    def _from_csv(self, data: bytes) -> str:
        text = data.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        return "\n".join(" | ".join(cell for cell in row) for row in rows)

    def _from_xlsx(self, data: bytes) -> str:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            lines = [f"# {sheet.title}"]
            for row in sheet.iter_rows(values_only=True):
                lines.append(" | ".join("" if cell is None else str(cell) for cell in row))
            sheets.append("\n".join(lines))
        workbook.close()
        return "\n\n".join(sheets)
